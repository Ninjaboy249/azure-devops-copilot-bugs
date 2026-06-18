"""Flask Web Application for Azure DevOps Bug Tracker with OAuth2 Authentication."""

import os
import io
import csv
import json
import base64
import secrets
import logging
from datetime import datetime
from functools import wraps
from pathlib import Path

import requests
from flask import (
    Flask, render_template, request, redirect, url_for,
    session, jsonify, send_file, flash
)
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", secrets.token_hex(32))

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# Azure DevOps OAuth Configuration
AZURE_DEVOPS_AUTH_URL = "https://app.vssps.visualstudio.com/oauth2/authorize"
AZURE_DEVOPS_TOKEN_URL = "https://app.vssps.visualstudio.com/oauth2/token"
AZURE_DEVOPS_APP_ID = os.getenv("AZURE_DEVOPS_APP_ID", "")
AZURE_DEVOPS_APP_SECRET = os.getenv("AZURE_DEVOPS_APP_SECRET", "")
AZURE_DEVOPS_CALLBACK_URL = os.getenv("AZURE_DEVOPS_CALLBACK_URL", "http://localhost:5000/callback")
AZURE_DEVOPS_SCOPE = "vso.work vso.project"

# PAT-based auth fallback
AZURE_DEVOPS_PAT = os.getenv("AZURE_DEVOPS_PAT", "")
AZURE_DEVOPS_ORG_URL = os.getenv("AZURE_DEVOPS_ORG_URL", "")
AZURE_DEVOPS_PROJECT = os.getenv("AZURE_DEVOPS_PROJECT", "")


def login_required(f):
    """Decorator to require authentication."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("authenticated"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function


def get_auth_headers():
    """Get authorization headers based on auth method."""
    if session.get("access_token"):
        return {"Authorization": f"Bearer {session['access_token']}", "Content-Type": "application/json"}
    elif session.get("pat"):
        credentials = base64.b64encode(f":{session['pat']}".encode()).decode()
        return {"Authorization": f"Basic {credentials}", "Content-Type": "application/json"}
    return {}


def execute_wiql(org_url, project, query):
    """Execute a WIQL query against Azure DevOps."""
    url = f"{org_url}/{project}/_apis/wit/wiql?api-version=7.1"
    headers = get_auth_headers()

    response = requests.post(url, json={"query": query}, headers=headers, timeout=30)
    response.raise_for_status()
    result = response.json()

    work_item_ids = [item["id"] for item in result.get("workItems", [])]
    if not work_item_ids:
        return []

    # Fetch work items in batches
    all_items = []
    for i in range(0, len(work_item_ids), 200):
        batch = work_item_ids[i:i + 200]
        ids_str = ",".join(str(id) for id in batch)
        wi_url = f"{org_url}/{project}/_apis/wit/workitems?ids={ids_str}&$expand=all&api-version=7.1"
        wi_response = requests.get(wi_url, headers=headers, timeout=30)
        wi_response.raise_for_status()
        all_items.extend(wi_response.json().get("value", []))

    return all_items


def parse_prompt_to_wiql(prompt, project):
    """Convert natural language prompt to WIQL query."""
    prompt_lower = prompt.lower().strip()

    # Base fields
    fields = """[System.Id], [System.Title], [System.State],
               [Microsoft.VSTS.Common.Priority], [System.AssignedTo],
               [System.CreatedDate], [Microsoft.VSTS.Common.Severity],
               [System.IterationPath], [System.AreaPath], [System.Tags]"""

    # Pattern matching for common prompts
    if any(kw in prompt_lower for kw in ["active bug", "open bug", "active defect"]):
        return f"SELECT {fields} FROM WorkItems WHERE [System.WorkItemType] = 'Bug' AND [System.State] = 'Active' ORDER BY [Microsoft.VSTS.Common.Priority] ASC"

    elif any(kw in prompt_lower for kw in ["my bug", "assigned to me", "my defect"]):
        return f"SELECT {fields} FROM WorkItems WHERE [System.WorkItemType] = 'Bug' AND [System.AssignedTo] = @Me ORDER BY [Microsoft.VSTS.Common.Priority] ASC"

    elif any(kw in prompt_lower for kw in ["p1", "critical", "blocker", "priority 1"]):
        if "p2" in prompt_lower or "priority 2" in prompt_lower:
            return f"SELECT {fields} FROM WorkItems WHERE [System.WorkItemType] = 'Bug' AND [Microsoft.VSTS.Common.Priority] <= 2 ORDER BY [Microsoft.VSTS.Common.Priority] ASC"
        return f"SELECT {fields} FROM WorkItems WHERE [System.WorkItemType] = 'Bug' AND [Microsoft.VSTS.Common.Priority] = 1 ORDER BY [System.CreatedDate] DESC"

    elif any(kw in prompt_lower for kw in ["p2", "high priority", "priority 2"]):
        return f"SELECT {fields} FROM WorkItems WHERE [System.WorkItemType] = 'Bug' AND [Microsoft.VSTS.Common.Priority] = 2 ORDER BY [System.CreatedDate] DESC"

    elif any(kw in prompt_lower for kw in ["sprint", "current iteration", "this sprint"]):
        return f"SELECT {fields} FROM WorkItems WHERE [System.WorkItemType] = 'Bug' AND [System.IterationPath] = @CurrentIteration ORDER BY [Microsoft.VSTS.Common.Priority] ASC"

    elif any(kw in prompt_lower for kw in ["resolved", "fixed"]):
        return f"SELECT {fields} FROM WorkItems WHERE [System.WorkItemType] = 'Bug' AND [System.State] = 'Resolved' ORDER BY [System.ChangedDate] DESC"

    elif any(kw in prompt_lower for kw in ["closed"]):
        return f"SELECT {fields} FROM WorkItems WHERE [System.WorkItemType] = 'Bug' AND [System.State] = 'Closed' ORDER BY [System.ChangedDate] DESC"

    elif any(kw in prompt_lower for kw in ["new bug", "newly created", "recent bug"]):
        return f"SELECT {fields} FROM WorkItems WHERE [System.WorkItemType] = 'Bug' AND [System.State] = 'New' ORDER BY [System.CreatedDate] DESC"

    elif any(kw in prompt_lower for kw in ["unassigned"]):
        return f"SELECT {fields} FROM WorkItems WHERE [System.WorkItemType] = 'Bug' AND [System.AssignedTo] = '' ORDER BY [Microsoft.VSTS.Common.Priority] ASC"

    elif any(kw in prompt_lower for kw in ["all bug", "every bug", "total bug"]):
        return f"SELECT {fields} FROM WorkItems WHERE [System.WorkItemType] = 'Bug' ORDER BY [Microsoft.VSTS.Common.Priority] ASC, [System.CreatedDate] DESC"

    # Default: all active bugs
    return f"SELECT {fields} FROM WorkItems WHERE [System.WorkItemType] = 'Bug' AND [System.State] <> 'Closed' ORDER BY [Microsoft.VSTS.Common.Priority] ASC"


def format_bugs(raw_bugs):
    """Format raw API response into clean bug data."""
    bugs = []
    for bug in raw_bugs:
        fields = bug.get("fields", {})
        assignee = fields.get("System.AssignedTo", {})
        assignee_name = assignee.get("displayName", "Unassigned") if isinstance(assignee, dict) else "Unassigned"

        bugs.append({
            "id": bug.get("id"),
            "title": fields.get("System.Title", ""),
            "state": fields.get("System.State", ""),
            "priority": fields.get("Microsoft.VSTS.Common.Priority", ""),
            "severity": fields.get("Microsoft.VSTS.Common.Severity", ""),
            "assigned_to": assignee_name,
            "created_date": fields.get("System.CreatedDate", "")[:10] if fields.get("System.CreatedDate") else "",
            "iteration_path": fields.get("System.IterationPath", ""),
            "area_path": fields.get("System.AreaPath", ""),
            "tags": fields.get("System.Tags", ""),
        })
    return bugs


# ==================== ROUTES ====================

@app.route("/")
def index():
    """Landing page."""
    if session.get("authenticated"):
        return redirect(url_for("dashboard"))
    return render_template("login.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    """Login page with PAT or OAuth options."""
    if request.method == "POST":
        auth_method = request.form.get("auth_method")

        if auth_method == "pat":
            pat = request.form.get("pat", "").strip()
            org_url = request.form.get("org_url", "").strip().rstrip("/")
            project = request.form.get("project", "").strip()

            if not all([pat, org_url, project]):
                flash("All fields are required.", "error")
                return render_template("login.html")

            # Verify credentials
            credentials = base64.b64encode(f":{pat}".encode()).decode()
            headers = {"Authorization": f"Basic {credentials}", "Content-Type": "application/json"}
            test_url = f"{org_url}/_apis/projects/{project}?api-version=7.1"

            try:
                resp = requests.get(test_url, headers=headers, timeout=10)
                if resp.status_code == 200:
                    session["authenticated"] = True
                    session["pat"] = pat
                    session["org_url"] = org_url
                    session["project"] = project
                    session["project_name"] = resp.json().get("name", project)
                    flash("Successfully connected to Azure DevOps!", "success")
                    return redirect(url_for("dashboard"))
                else:
                    flash(f"Authentication failed (HTTP {resp.status_code}). Check your credentials.", "error")
            except requests.exceptions.RequestException as e:
                flash(f"Connection error: {str(e)}", "error")

        elif auth_method == "oauth":
            # Initiate OAuth flow
            state = secrets.token_urlsafe(32)
            session["oauth_state"] = state
            auth_url = (
                f"{AZURE_DEVOPS_AUTH_URL}?client_id={AZURE_DEVOPS_APP_ID}"
                f"&response_type=Assertion&state={state}"
                f"&scope={AZURE_DEVOPS_SCOPE}"
                f"&redirect_uri={AZURE_DEVOPS_CALLBACK_URL}"
            )
            return redirect(auth_url)

    return render_template("login.html")


@app.route("/callback")
def oauth_callback():
    """OAuth2 callback handler."""
    code = request.args.get("code")
    state = request.args.get("state")

    if state != session.get("oauth_state"):
        flash("Invalid OAuth state. Please try again.", "error")
        return redirect(url_for("login"))

    # Exchange code for token
    token_data = {
        "client_assertion_type": "urn:ietf:params:oauth:client-assertion-type:jwt-bearer",
        "client_assertion": AZURE_DEVOPS_APP_SECRET,
        "grant_type": "urn:ietf:params:oauth:grant-type:jwt-bearer",
        "assertion": code,
        "redirect_uri": AZURE_DEVOPS_CALLBACK_URL,
    }

    try:
        resp = requests.post(AZURE_DEVOPS_TOKEN_URL, data=token_data, timeout=10)
        if resp.status_code == 200:
            token_info = resp.json()
            session["authenticated"] = True
            session["access_token"] = token_info["access_token"]
            session["org_url"] = request.form.get("org_url", AZURE_DEVOPS_ORG_URL)
            session["project"] = request.form.get("project", AZURE_DEVOPS_PROJECT)
            flash("Successfully authenticated via OAuth!", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("OAuth token exchange failed.", "error")
    except Exception as e:
        flash(f"OAuth error: {str(e)}", "error")

    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    """Main dashboard with prompt interface."""
    return render_template("dashboard.html", project=session.get("project_name", session.get("project")))


@app.route("/query", methods=["POST"])
@login_required
def query_bugs():
    """Process user prompt and fetch bugs."""
    data = request.get_json()
    prompt = data.get("prompt", "").strip()

    if not prompt:
        return jsonify({"error": "Please enter a prompt."}), 400

    org_url = session.get("org_url")
    project = session.get("project")

    try:
        wiql = parse_prompt_to_wiql(prompt, project)
        raw_bugs = execute_wiql(org_url, project, wiql)
        bugs = format_bugs(raw_bugs)

        # Store in session for export
        session["last_results"] = bugs
        session["last_query"] = prompt

        return jsonify({
            "success": True,
            "bugs": bugs,
            "total": len(bugs),
            "query": prompt,
            "wiql": wiql,
        })

    except requests.exceptions.HTTPError as e:
        logger.error(f"API error: {e}")
        return jsonify({"error": f"Azure DevOps API error: {e.response.status_code}"}), 500
    except Exception as e:
        logger.error(f"Query error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/export/<format_type>", methods=["POST"])
@login_required
def export_data(format_type):
    """Export bug data in CSV or Excel format."""
    data = request.get_json()
    bugs = data.get("bugs", [])

    if not bugs:
        return jsonify({"error": "No data to export."}), 400

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if format_type == "csv":
        output = io.StringIO()
        if bugs:
            writer = csv.DictWriter(output, fieldnames=bugs[0].keys())
            writer.writeheader()
            writer.writerows(bugs)

        mem = io.BytesIO()
        mem.write(output.getvalue().encode("utf-8"))
        mem.seek(0)

        return send_file(
            mem,
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"bugs_export_{timestamp}.csv",
        )

    elif format_type == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Bugs"

            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")

            if bugs:
                headers = list(bugs[0].keys())
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                priority_fills = {
                    1: PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
                    2: PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),
                    3: PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
                }

                for row_idx, bug in enumerate(bugs, 2):
                    for col_idx, key in enumerate(headers, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=bug.get(key, ""))
                        if key == "priority" and bug.get(key) in priority_fills:
                            cell.fill = priority_fills[bug[key]]
                            cell.font = Font(bold=True, color="FFFFFF")

                # Auto-fit columns
                for col_idx in range(1, len(headers) + 1):
                    from openpyxl.utils import get_column_letter
                    max_len = max(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

            mem = io.BytesIO()
            wb.save(mem)
            mem.seek(0)

            return send_file(
                mem,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=f"bugs_export_{timestamp}.xlsx",
            )

        except ImportError:
            return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    return jsonify({"error": "Invalid format. Use 'csv' or 'excel'."}), 400


@app.route("/logout")
def logout():
    """Clear session and logout."""
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("index"))


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
