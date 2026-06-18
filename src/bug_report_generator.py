"""Bug Report Generator - CSV, Excel, HTML, and JSON outputs."""

import csv
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


class BugReportGenerator:
    """Generates bug reports in multiple formats."""

    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _extract_bug_fields(self, bugs: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """Extract relevant fields from raw work item data."""
        extracted = []
        for bug in bugs:
            fields = bug.get("fields", {})
            assignee = fields.get("System.AssignedTo", {})
            assignee_name = assignee.get("displayName", "Unassigned") if isinstance(assignee, dict) else "Unassigned"

            extracted.append({
                "ID": bug.get("id"),
                "Title": fields.get("System.Title", ""),
                "State": fields.get("System.State", ""),
                "Priority": fields.get("Microsoft.VSTS.Common.Priority", ""),
                "Severity": fields.get("Microsoft.VSTS.Common.Severity", ""),
                "Assigned To": assignee_name,
                "Created Date": fields.get("System.CreatedDate", ""),
                "Changed Date": fields.get("System.ChangedDate", ""),
                "Iteration Path": fields.get("System.IterationPath", ""),
                "Area Path": fields.get("System.AreaPath", ""),
                "Tags": fields.get("System.Tags", ""),
                "URL": bug.get("url", ""),
            })
        return extracted

    def to_csv(self, bugs: list[dict[str, Any]], filename: str = "bugs_report.csv") -> Path:
        """Export bugs to CSV file."""
        filepath = self.output_dir / filename
        extracted = self._extract_bug_fields(bugs)

        if not extracted:
            logger.warning("No bugs to export to CSV")
            return filepath

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=extracted[0].keys())
            writer.writeheader()
            writer.writerows(extracted)

        logger.info(f"CSV report saved: {filepath} ({len(extracted)} bugs)")
        return filepath

    def to_excel(self, bugs: list[dict[str, Any]], filename: str = "bugs_report.xlsx") -> Path:
        """Export bugs to Excel file."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            raise

        filepath = self.output_dir / filename
        extracted = self._extract_bug_fields(bugs)

        if not extracted:
            logger.warning("No bugs to export to Excel")
            return filepath

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bug Report"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")

        # Write headers
        headers = list(extracted[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Priority color mapping
        priority_fills = {
            1: PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            2: PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),
            3: PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
            4: PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
        }

        # Write data
        for row_idx, bug in enumerate(extracted, 2):
            for col_idx, key in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=bug[key])
                if key == "Priority" and bug[key] in priority_fills:
                    cell.fill = priority_fills[bug[key]]
                    cell.font = Font(bold=True, color="FFFFFF" if bug[key] <= 2 else "000000")

        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            max_length = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

        # Add summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.cell(row=1, column=1, value="Bug Summary Report").font = Font(bold=True, size=14)
        ws_summary.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws_summary.cell(row=3, column=1, value=f"Total Bugs: {len(extracted)}")

        # Priority breakdown
        ws_summary.cell(row=5, column=1, value="Priority Breakdown").font = Font(bold=True)
        priority_counts = {}
        for bug in extracted:
            p = f"P{bug['Priority']}" if bug['Priority'] else "Unset"
            priority_counts[p] = priority_counts.get(p, 0) + 1
        for idx, (priority, count) in enumerate(sorted(priority_counts.items()), 6):
            ws_summary.cell(row=idx, column=1, value=priority)
            ws_summary.cell(row=idx, column=2, value=count)

        wb.save(filepath)
        logger.info(f"Excel report saved: {filepath} ({len(extracted)} bugs)")
        return filepath

    def to_json(self, bugs: list[dict[str, Any]], filename: str = "bugs_report.json") -> Path:
        """Export bugs to JSON file."""
        filepath = self.output_dir / filename
        extracted = self._extract_bug_fields(bugs)

        report = {
            "metadata": {
                "generated_at": datetime.now().isoformat(),
                "total_bugs": len(extracted),
            },
            "bugs": extracted,
        }

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(report, f, indent=2, default=str)

        logger.info(f"JSON report saved: {filepath} ({len(extracted)} bugs)")
        return filepath

    def to_html(self, bugs: list[dict[str, Any]], filename: str = "bugs_dashboard.html") -> Path:
        """Export bugs to an HTML dashboard."""
        filepath = self.output_dir / filename
        extracted = self._extract_bug_fields(bugs)

        # Calculate statistics
        total = len(extracted)
        active = sum(1 for b in extracted if b["State"] == "Active")
        resolved = sum(1 for b in extracted if b["State"] == "Resolved")
        closed = sum(1 for b in extracted if b["State"] == "Closed")
        p1_count = sum(1 for b in extracted if b["Priority"] == 1)
        p2_count = sum(1 for b in extracted if b["Priority"] == 2)

        # Priority distribution for chart
        priority_counts = {}
        for bug in extracted:
            p = f"P{bug['Priority']}" if bug['Priority'] else "Unset"
            priority_counts[p] = priority_counts.get(p, 0) + 1

        # Build HTML
        priority_colors = {"P1": "#dc3545", "P2": "#fd7e14", "P3": "#ffc107", "P4": "#28a745", "Unset": "#6c757d"}
        rows_html = ""
        for bug in extracted:
            priority_class = f"priority-{bug['Priority']}" if bug['Priority'] else "priority-none"
            rows_html += f"""
            <tr class="{priority_class}">
                <td>{bug['ID']}</td>
                <td>{bug['Title']}</td>
                <td><span class="badge state-{bug['State'].lower()}">{bug['State']}</span></td>
                <td><span class="badge priority-badge-{bug['Priority']}">P{bug['Priority']}</span></td>
                <td>{bug['Severity']}</td>
                <td>{bug['Assigned To']}</td>
                <td>{bug['Iteration Path']}</td>
                <td>{bug['Tags']}</td>
            </tr>"""

        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Azure DevOps Bug Dashboard</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f4f6f8; color: #333; padding: 20px; }}
        .container {{ max-width: 1400px; margin: 0 auto; }}
        h1 {{ color: #0078d4; margin-bottom: 5px; }}
        .subtitle {{ color: #666; margin-bottom: 20px; }}
        .stats-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 15px; margin-bottom: 30px; }}
        .stat-card {{ background: white; border-radius: 8px; padding: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); text-align: center; }}
        .stat-card .number {{ font-size: 2.5em; font-weight: bold; }}
        .stat-card .label {{ color: #666; margin-top: 5px; }}
        .stat-card.total .number {{ color: #0078d4; }}
        .stat-card.active .number {{ color: #dc3545; }}
        .stat-card.resolved .number {{ color: #28a745; }}
        .stat-card.closed .number {{ color: #6c757d; }}
        .stat-card.p1 .number {{ color: #dc3545; }}
        .stat-card.p2 .number {{ color: #fd7e14; }}
        table {{ width: 100%; border-collapse: collapse; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        th {{ background: #0078d4; color: white; padding: 12px 15px; text-align: left; font-weight: 600; }}
        td {{ padding: 10px 15px; border-bottom: 1px solid #eee; }}
        tr:hover {{ background: #f0f7ff; }}
        .badge {{ padding: 3px 8px; border-radius: 12px; font-size: 0.85em; font-weight: 500; }}
        .state-active {{ background: #ffeeba; color: #856404; }}
        .state-resolved {{ background: #c3e6cb; color: #155724; }}
        .state-closed {{ background: #d6d8db; color: #383d41; }}
        .state-new {{ background: #b8daff; color: #004085; }}
        .priority-badge-1 {{ background: #dc3545; color: white; }}
        .priority-badge-2 {{ background: #fd7e14; color: white; }}
        .priority-badge-3 {{ background: #ffc107; color: #333; }}
        .priority-badge-4 {{ background: #28a745; color: white; }}
        .priority-1 {{ border-left: 4px solid #dc3545; }}
        .priority-2 {{ border-left: 4px solid #fd7e14; }}
        .footer {{ text-align: center; margin-top: 20px; color: #999; font-size: 0.85em; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🐛 Azure DevOps Bug Dashboard</h1>
        <p class="subtitle">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>

        <div class="stats-grid">
            <div class="stat-card total"><div class="number">{total}</div><div class="label">Total Bugs</div></div>
            <div class="stat-card active"><div class="number">{active}</div><div class="label">Active</div></div>
            <div class="stat-card resolved"><div class="number">{resolved}</div><div class="label">Resolved</div></div>
            <div class="stat-card closed"><div class="number">{closed}</div><div class="label">Closed</div></div>
            <div class="stat-card p1"><div class="number">{p1_count}</div><div class="label">P1 Critical</div></div>
            <div class="stat-card p2"><div class="number">{p2_count}</div><div class="label">P2 High</div></div>
        </div>

        <table>
            <thead>
                <tr>
                    <th>ID</th><th>Title</th><th>State</th><th>Priority</th>
                    <th>Severity</th><th>Assigned To</th><th>Sprint</th><th>Tags</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>

        <div class="footer">
            <p>Azure DevOps Bug Dashboard | Powered by GitHub Copilot Agent + MCP</p>
        </div>
    </div>
</body>
</html>"""

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html_content)

        logger.info(f"HTML dashboard saved: {filepath} ({total} bugs)")
        return filepath

    def generate_all_reports(
        self, bugs: list[dict[str, Any]], prefix: str = "bugs"
    ) -> dict[str, Path]:
        """Generate all report formats."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        reports = {}

        reports["csv"] = self.to_csv(bugs, f"{prefix}_{timestamp}.csv")
        reports["json"] = self.to_json(bugs, f"{prefix}_{timestamp}.json")
        reports["html"] = self.to_html(bugs, f"{prefix}_{timestamp}.html")

        try:
            reports["excel"] = self.to_excel(bugs, f"{prefix}_{timestamp}.xlsx")
        except ImportError:
            logger.warning("Skipping Excel export - openpyxl not installed")

        return reports
